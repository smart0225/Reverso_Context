import openpyxl
import pandas as pd
import psutil

#使用pandas新增一個EXCEL
def rw_xlsx(str,xls_file,dda,sName):
    write = pd.ExcelWriter(xls_file)
    dda.to_excel(write,sheet_name=sName)
    print(sName)

#建立xlsx檔案
def buildxlsx(newfile):
    # 建立 Excel 活頁簿
    wb = openpyxl.Workbook()
    #wb._named_styles['Normal'].number_format = '#,##0.00'
    # 取得作用中的工作表
    ws = wb.active
    # 設定工作表名稱
    ws.title = "(上市櫃)借劵賣出餘額增加"
    ws1 = wb.create_sheet("(上市櫃)借劵賣出餘額減少")
    ws2 = wb.create_sheet("(上櫃)借券賣出餘額增加")
    ws3 = wb.create_sheet("(上櫃)借券賣出餘額減少")
    ws4 = wb.create_sheet("(上市)借券賣出餘額增加")
    # 儲存 Excel 活頁簿至檔案
    wb.save(filename=newfile)
    print(newfile)

#kill excel及chrome處理程序
def Kill_XLSX():
    for proc in psutil.process_iter():
        if proc.name() == "chrome.exe" or proc.name() == "EXCEL.EXE":
            proc.kill()

#修改EXCEL儲存格
def Fix_Cells(xls_file):
    # 開啟EXCEL檔案
    wb = openpyxl.load_workbook(xls_file)
    # Sheets Count數
    res = len(wb.sheetnames)
    # While迴圈Sheets Count
    count = 0
    while count < res:
        wb.active = count
        ws = wb.active
        #FOR迴圈 確認每列每欄Cell
        for row in ws:
            a = 0
            for cell in row:
                a += 1
                try:
                    #將","符號取代刪除
                    cell.value = cell.value.replace(',', '')
                    #針對第一及第三欄位變更成整數及浮點數
                    if (a == 1 or a == 3):
                        cell.value = int(cell.value)
                    else:
                        cell.value = float(cell.value)
                except:  # 當程式出現異常時執行這邊的程式碼
                    #修改整數及浮點數失敗表示為"文字"，維持原狀態 Print出來
                    print(cell.value)

        count += 1
    #將修正後最後結果儲存
    wb.save(xls_file)